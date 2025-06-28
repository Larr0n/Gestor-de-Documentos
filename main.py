#librerias y funciones
import ttkbootstrap as tb
import tkinter as tk
from ttkbootstrap.constants import *
from tkinter import ttk, filedialog, messagebox
from tksheet import Sheet
import openpyxl
from openpyxl.utils import get_column_letter
from db.conexion import conexion
from db.consultas import consulta_default, consulta_filtrada
from funciones import guardar_datos, insertar_fila, eliminar_fila, configurar_logger, buckup
import logging
from PIL import Image, ImageTk
from tkinter import ttk
import sys 
import os
import bcrypt

configurar_logger()

#ventana principal
root = tb.Window(themename="flatly")
buckup()
root.withdraw()
root.title("Gestor de Documentos")
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  
    except Exception:
        base_path = os.path.abspath(".")  

    return os.path.join(base_path, relative_path)
icon_path = resource_path('img/folder.ico')
root.iconbitmap(icon_path)
def centrar_ventana(ventana, ancho, alto):
    ventana.update_idletasks() 
    pantalla_ancho = ventana.winfo_screenwidth()
    pantalla_alto = ventana.winfo_screenheight()
    x = int((pantalla_ancho / 2) - (ancho / 2))
    y = int((pantalla_alto / 2) - (alto / 2))
    ventana.geometry(f"{ancho}x{alto}+{x}+{y}")
centrar_ventana(root, 1290, 550)
root.resizable(True, True)
base = None
ids_filas = []




#splash window para estética
def mostrar_splash(root):
    splash = tk.Toplevel(root)
    splash.overrideredirect(True)
    centrar_ventana(splash, 530, 380)

    img_path = resource_path("img/folder.png")
    logo = Image.open(img_path)
    logo = logo.resize((110, 110))
    logo_img = ImageTk.PhotoImage(logo)
    logo_label = tk.Label(splash, image=logo_img)
    logo_label.image = logo_img
    logo_label.pack(pady=(25, 10))

    # Títulos
    tk.Label(splash, text="Gestor de Documentos", font=("Segoe UI", 16, "bold"), bg="#f2f4f8", fg="#2c3e50").pack()

    # Texto de carga
    tk.Label(splash, text="Cargando sistema...", font=("Segoe UI", 10, "italic"), bg="#f2f4f8", fg="#34495e").pack(pady=(15, 5))

    # Barra de progreso animada
    style = ttk.Style(splash)
    style.configure("TProgressbar", thickness=8, troughcolor='#000000', background='#0984e3', bordercolor='#000000')
    progress = ttk.Progressbar(splash, mode="indeterminate", length=400, style="TProgressbar")
    progress.pack(pady=(10, 20))
    progress.start(60)  

    # Desarrollador
    tk.Label(splash, text="Desarrollada por Pedro Larrondo", font=("Segoe UI", 9, "italic"), bg="#f2f4f8", fg="#7f8c8d").pack(side="bottom", pady=20)

    def continuar():
        progress.stop()
        splash.destroy()
        root.deiconify()
        main(root)

    splash.after(6000, continuar)


#frame principal
main_frame = tb.Frame(root, padding=15)
main_frame.pack(fill="both", expand=True)


#filtros
top_frame = tb.LabelFrame(main_frame, text="Filtros de Búsqueda", padding=10)
top_frame.pack(fill="x", pady=(0, 10))

#variables de filtros
dni_var = tk.StringVar()
apellido_var = tk.StringVar()
nombre_var = tk.StringVar()
num_var = tk.StringVar()
grado_var = tk.StringVar()
tipo_var = tk.StringVar()
num_filas_var = tk.StringVar()

#jerarquias y tipos de legajo
grados = [
    "Grado Primero", "Grado Segundo", "Grado Tercero"
]

tipos = [
    "Legajo Personal", "Legajo de Servicios", "Legajo de Salud", "Legajo de Familia",
    "Expediente de Retiro", "Expediente Médico",
]

cantidad = [ 25, 50, 75, 100, "TODAS"]

#lista de filtros con etiquetas, variables y tipo de entrada
filtros = [
    ("DNI", dni_var, "entry"), ("Apellido", apellido_var, "entry"), ("Nombre", nombre_var, "entry"),
    ("Número de Legajo", num_var, "entry"), ("Grado", grado_var, "combo", grados), ("OC", tipo_var, "combo", tipos),
    ("Número de Filas", num_filas_var, "combo", cantidad)
]

for i, (texto, var, tipo, *opciones) in enumerate(filtros):
    fila = 0 if i < 5 else 1  
    columna = i*2 if i < 5 else (i - 5)*2  

    tb.Label(top_frame, text=f"{texto}:", font=("Segoe UI", 9, "bold")).grid(row=fila, column=columna, sticky="e", padx=5, pady=5)

    if tipo == "entry":
        tb.Entry(top_frame, textvariable=var, font=("Segoe UI", 9), width=12).grid(row=fila, column=columna+1, padx=10, pady=5)
    else:
        estado = "normal" if texto == "Número de Filas" else "readonly"
        ttk.Combobox(top_frame, textvariable=var, values=opciones[0], state=estado, width=15).grid(row=fila, column=columna+1, padx=10, pady=5)

cantidad_var = tb.StringVar()
cantidad_var.set("Resultados encontrados: 0")

# realizar busquedas
def ejecutar_busqueda():
    global base, ids_filas
    filtros = {
        "dni": dni_var.get().strip(),
        "apellido": apellido_var.get().strip().lower(),
        "nombre": nombre_var.get().strip().lower(),
        "num_legajo": num_var.get().strip().lower(),
        "grado": grado_var.get().strip().lower(),
        "tipo": tipo_var.get().strip().lower(),
        "num_fila": num_filas_var.get().strip().lower(),
    }
    busqueda = consulta_filtrada(filtros)
    if busqueda:
        ids_filas = [fila[0] for fila in busqueda]
        base.set_sheet_data([list(fila[1:]) for fila in busqueda])
    else:
        base.set_sheet_data([])
        ids_filas = []
    cantidad_var.set(f"Cantidad de resultados: {len(ids_filas)}")

# limpiar las busquedas
def limpiar_filtro():
    for var in [dni_var, apellido_var, nombre_var, num_var, grado_var, tipo_var, num_filas_var]:
        var.set("")
    busqueda = consulta_default(conexion())
    if busqueda:
        ids_filas[:] = [fila[0] for fila in busqueda]
        base.set_sheet_data([list(fila[1:]) for fila in busqueda])
    cantidad_var.set(f"Cantidad de resultados: {len(ids_filas)}")

# botones bsucar y limpiar
btn_frame = tb.Frame(top_frame)
btn_frame.grid(row=0, column=len(filtros)*2, padx=(10, 0))
tb.Button(btn_frame, text="🔍 Buscar", command=ejecutar_busqueda, bootstyle="primary").pack(side="left", padx=20)
tb.Button(btn_frame, text="🧹 Limpiar", command=limpiar_filtro, bootstyle="primary").pack(side="left", padx=20)


#funcion princiapl
def main(root):
    global base, ids_filas

    con = conexion()
    if not con:
        print("No se pudo conectar a la base de datos")
        return

    datos = consulta_default(con)

    ids_filas = [fila[0] for fila in datos]
    base_datos = [list(fila[1:]) for fila in datos]
    cantidad_var.set(f"Cantidad de resultados: {len(ids_filas)}")

    #tabla excel con datos
    global base
    base = Sheet(
        main_frame,
        data=base_datos,
        headers=[
            'N° DE LEGAJO', '            GRADO            ', '       DNI       ', '            APELLIDOS            ', '              NOMBRES              ',
            '     FECHA DE RETIRO     ',
            '           TIPO DE DOCUMENTO (OC)           ', 
            '      ENVIADO A ?      ', ' FECHA DEVOLUCIÓN'
        ],
        show_x_scrollbar=True,
        show_y_scrollbar=True,
        show_row_index=True
    )
    base.pack(expand=True, fill="both", pady=(0, 15))
    #funcionalidades permitidas
    base.enable_bindings((
        "single_select", "row_select", "row_index", "arrowkeys",
        "right_click_popup_menu", "rc_select", "copy", "cut", "paste",
        "delete", "undo", "edit_cell"
    ))
    base.set_options(selection_mode="row")
    base.extra_bindings("end_edit_cell", lambda e: guardar_datos(e, base, ids_filas))
    base.set_all_column_widths()
    base.column_width(6, 210)

    #botones de abajo
    acciones = tb.Frame(main_frame)
    acciones.pack(pady=10)
    etiqueta_resultados = tb.Label(acciones, textvariable=cantidad_var, font=("Arial", 10))
    etiqueta_resultados.pack(side="left", padx=20)
    etiqueta_resultados.config(text=f"Resultados encontrados: {cantidad}")
    tb.Button(acciones, text="➕ Agregar Fila", command=agregar_fila, bootstyle="info").pack(side="left", padx=10)
    tb.Button(acciones, text="🗑️ Eliminar Fila", command=lambda: eliminar_fila(base, ids_filas), bootstyle="danger").pack(side="left", padx=10)
    tb.Button(acciones, text="📤 Exportar a Excel", command=exportar_a_excel, bootstyle="success").pack(side="left", padx=10)

#agregar fila
def agregar_fila():
    try:
        nueva_fila = insertar_fila()
        if nueva_fila:
            ids_filas.append(nueva_fila[0])
            base.insert_row(list(nueva_fila[1:]))
            base.set_all_column_widths()
    except Exception as e:
        print("❌ Error al insertar fila:", e)

#con esto podemos exportar a excel los valores que aparecen en la pantalla (ya sean bsuquedas personalizadas o por default)
def exportar_a_excel():
    try:
        datos = base.get_sheet_data()
        encabezados = base.headers()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gestor de Documentos"
        for col_num, header in enumerate(encabezados, 1):
            ws.cell(row=1, column=col_num, value=header)
        for row_num, fila in enumerate(datos, 2):
            for col_num, celda in enumerate(fila, 1):
                ws.cell(row=row_num, column=col_num, value=celda)
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar como"
        )
        if archivo:
            wb.save(archivo)
            messagebox.showinfo("Exportación exitosa", f"Búsqueda exportada a:\n{archivo}")
            logging.info(f"Se ha realizado una exportación a {archivo}")
    except Exception as e:
        messagebox.showerror("Error al exportar", f"Ocurrió un error:\n{e}")


#mostrar ventana de login, validar usuarios, redigirir a splash screen y luego a página principal
def mostrar_login(root):
    def limpiar_placeholder(event, entry, placeholder, is_password=False):
        if entry.get() == placeholder:
            entry.delete(0, tk.END)
            entry.config(fg="black", show="*" if is_password else "")

    def agregar_placeholder(event, entry, placeholder, is_password=False):
        if not entry.get():
            entry.insert(0, placeholder)
            entry.config(fg="grey", show="")

    root.withdraw()
    login_window = tk.Toplevel()
    login_window.title("Login")
    login_window.geometry("260x350")
    centrar_ventana(login_window, 270, 370)
    login_window.grab_set()

    img_path = resource_path("img/usuario.png")
    img = Image.open(img_path) 
    img = img.resize((100, 120))
    photo = ImageTk.PhotoImage(img)
    logo = tk.Label(login_window, image=photo)
    logo.image = photo
    logo.pack(pady=(20, 10))

    # Entrada Usuario
    usuario_entry = tk.Entry(login_window, font=("Segoe UI", 10), fg="grey", justify="center", relief="solid", bd=1)
    usuario_entry.insert(0, "Usuario")
    usuario_entry.bind("<FocusIn>", lambda e: limpiar_placeholder(e, usuario_entry, "Usuario"))
    usuario_entry.bind("<FocusOut>", lambda e: agregar_placeholder(e, usuario_entry, "Usuario"))
    usuario_entry.pack(pady=8, ipadx=8, ipady=5)

    # Entrada Contraseña
    contrasena_entry = tk.Entry(login_window, font=("Segoe UI", 10), fg="grey", justify="center", relief="solid", bd=1)
    contrasena_entry.insert(0, "Contraseña")
    contrasena_entry.bind("<FocusIn>", lambda e: limpiar_placeholder(e, contrasena_entry, "Contraseña", True))
    contrasena_entry.bind("<FocusOut>", lambda e: agregar_placeholder(e, contrasena_entry, "Contraseña", True))
    contrasena_entry.pack(pady=8, ipadx=8, ipady=5)

    # Botón Ingresar
    tk.Button(
        login_window, text="Ingresar", command=lambda: verificar_credenciales(),
        bg="#4a90e2", fg="white", font=("Segoe UI", 10, "bold"),
        relief="flat", padx=10, pady=5
    ).pack(pady=15)

    #verificamos dentro de la base de datos que el usuario y contraseña sean correctos
    def verificar_credenciales():
        usuario = usuario_entry.get()
        contrasena = contrasena_entry.get()

        con = conexion()
        if con is None:
            messagebox.showerror("Error", "❌ No se pudo conectar a la base de datos")
            return

        cur = con.cursor()
        cur.execute("SELECT password FROM usuarios WHERE username = %s", (usuario,))
        resultado = cur.fetchone()
        cur.close()
        con.close()

        if resultado and bcrypt.checkpw(contrasena.encode('utf-8'), resultado[0].encode('utf-8')):
            login_window.destroy()
            logging.info(f"Ha ingresado el usuario 👤 {usuario}")
            
            mostrar_splash(root)
        else:
            messagebox.showerror("Error", "❌ Usuario o contraseña incorrectos")
            logging.info(f"❌ Se intentó entrar con el nombre de usuario {usuario} pero algo ha fallado")
            usuario_entry.focus()

#ejecutar programa
if __name__ == "__main__":
    mostrar_login(root)
    root.mainloop()